"""
make_excel_workbook.py -- build the Excel companion (excel/spacex_valuation.xlsx).

A formula-based implementation of the paper's valuation for readers who prefer Excel to Python:
the three-segment DCF laid out as a classic year-by-year cash-flow waterfall with a step-by-step
terminal value, the expansion claims by their closed forms, the Mars probability tree, a live
Monte Carlo of the correlated segment model (1,000 trials on _xlfn.NORM.S.INV(RAND()); press F9 to
redraw), the sum of parts, and the inversion (a value-versus-WACC grid with interpolation, no
macros). The xAI abandonment option enters as an imported value: it is an optimal-stopping
problem solved by least-squares Monte Carlo (venture_option.py) with no formula representation.

Conventions (standard financial-model colors): blue on yellow = inputs to change, black =
formulas, green = links across tabs, grey italic = the Python pipeline's values quoted for
checking. Any divergence between a formula and its grey neighbor beyond rounding or stated
Monte-Carlo error signals a problem.
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter as col

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent))
import numpy as np
from spacex_realoptions import FirmParams, simulate

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "excel" / "spacex_valuation.xlsx"

# the paper's stochastic distribution (200,000 fixed-seed paths), computed fresh for the
# grey check cells -- never hand-typed
_EQ = simulate(FirmParams(), n=200_000, seed=2026, stochastic=True)["equity"]
CHK = {"mean": float(_EQ.mean()), "p5": float(np.quantile(_EQ, 0.05)),
       "med": float(np.quantile(_EQ, 0.5)), "p95": float(np.quantile(_EQ, 0.95)),
       "pgt": float((_EQ > 1.8e6).mean())}

DEC = json.loads((ROOT / "output" / "tables" / "decomposition.json").read_text())
OP = DEC["option_params"]

BLUE = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10)
GREEN = Font(name="Arial", size=10, color="008000")
GREY = Font(name="Arial", size=9, italic=True, color="808080")
BOLD = Font(name="Arial", size=10, bold=True)
H1 = Font(name="Arial", size=13, bold=True)
H2 = Font(name="Arial", size=11, bold=True)
YELLOW = PatternFill("solid", start_color="FFF2CC")
M0 = "#,##0;(#,##0);-"
PCT = "0.0%"
PCT2 = "0.00%"

T = 11          # years 2026..2036
NTRIALS = 1000  # Monte Carlo trials


def fix_axes(chart, yfmt="#,##0", xfmt=None):
    """openpyxl charts omit axis tick labels unless these are set explicitly."""
    for ax, fmtv in ((chart.y_axis, yfmt), (chart.x_axis, xfmt)):
        ax.delete = False
        ax.tickLblPos = "nextTo"
        if fmtv:
            ax.number_format = fmtv
            ax.sourceLinked = False


def w(ws, r, c, value, font=BLACK, fmt=None, fill=None, bold=False):
    if isinstance(value, str) and value.startswith("=") and " " in value[:3]:
        raise ValueError(f"text cell would be parsed as a formula: {value!r}")
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = BOLD if bold else font
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = fill
    return cell


IN: dict[str, str] = {}        # input label -> absolute reference
DCFREF: dict[str, object] = {}
OPTREF: dict[str, str] = {}

SEGS = [("Launch", 4086, 40000, 0.30, 0.45, 3.5, 0.18, 0.08),
        ("Starlink", 11387, 120000, 0.35, 0.60, 4.0, 0.22, 0.10),
        ("xAI", 3201, 160000, -0.30, 0.25, 2.0, 0.40, 0.15)]


# ----------------------------------------------------------------------------------
def build_inputs(ws):
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 13
    ws.column_dimensions["C"].width = 95
    w(ws, 1, 1, "Inputs — change the blue cells; every tab updates", font=H1)
    w(ws, 2, 1, "Values in $M unless noted. Paper: ssrn.com/abstract=6918120 (Table 1, Appendix A).",
      font=GREY)
    r = 4

    def inp(label, value, source, fmt=M0):
        nonlocal r
        w(ws, r, 1, label)
        w(ws, r, 2, value, font=BLUE, fmt=fmt, fill=YELLOW)
        w(ws, r, 3, source, font=GREY)
        IN[label] = f"Inputs!$B${r}"
        r += 1

    def head(txt):
        nonlocal r
        r += 1
        w(ws, r, 1, txt, font=H2)
        r += 1

    head("Segments (2025 figures from the prospectus; targets, margins, volatilities per paper Table 1)")
    for name, rev0, tgt, m0, mT, s2c, v0, vbar in SEGS:
        inp(f"{name}: revenue 2025", rev0, "Prospectus, segment results (FWP acc. 0001628280-26-041013)")
        inp(f"{name}: revenue target 2036", tgt, "Damodaran, June 2026 valuation (paper Sec. 4)")
        msrc = {"Launch": "Development-adjusted; accounting segment margin was -16% after $3.0B Starship R&D (prospectus)",
                "Starlink": "Development-adjusted; accounting segment margin +39% (prospectus)",
                "xAI": "Development-adjusted; accounting segment margin about -198% (prospectus). Weighted, the three match Damodaran's R&D-capitalized 21.5% consolidated start"}[name]
        inp(f"{name}: operating margin 2025", m0, msrc, PCT)
        inp(f"{name}: terminal margin", mT, "Comparables-disciplined (paper Sec. 4); xAI 25% is the boundary case", PCT)
        inp(f"{name}: sales-to-capital", s2c, "Damodaran convention (paper Table 1)", "0.0")
        inp(f"{name}: revenue volatility, initial", v0, "Calibrated so dispersion lies between the two published simulations (Table 1)", PCT)
        inp(f"{name}: revenue volatility, long run", vbar, "Paper Table 1", PCT)

    head("Firm-level")
    inp("Cost of capital (WACC)", 0.0825, "Damodaran bottom-up CAPM: business-weighted relevered betas (paper Sec. 4.2)", PCT2)
    inp("Terminal growth", 0.0456, "Equals the June 2026 risk-free rate (Damodaran convention)", PCT2)
    inp("Terminal ROIC", 0.15, "Paper Table 1; sets perpetuity reinvestment g/ROIC", PCT)
    inp("Tax rate, current effective", 0.10, "Prospectus financials", PCT)
    inp("Tax rate, marginal", 0.25, "Paper Table 1", PCT)
    inp("Cash", 24747, "Prospectus balance sheet")
    inp("Debt", 22896, "Prospectus balance sheet")
    inp("IPO proceeds", 75000, "Offering: 555.6M shares at $135")
    inp("Shares outstanding (M)", 13091, "12,535M pre-offering + 556M new; CEO unvested award excluded (paper fn.)")
    inp("Common-factor correlation rho", 0.30, "Paper Table 1 (segments share one demand factor)", PCT)
    inp("Volatility decay kappa", 0.20, "Paper Appendix A.1 (uncertainty resolves over time)", "0.00")

    head("Expansion claims (paper Sec. 5.1, Appendix A.2)")
    inp("DTC: 2025 revenue", 500, "DTC underlying segment (Appendix A.2)")
    inp("DTC: revenue target 2036", 15000, "Between D2D forecasts of roughly $3-12B (2030) and ~$100B long-run")
    inp("DTC: margin 2025", -0.20, "Appendix A.2", PCT)
    inp("DTC: terminal margin", 0.55, "Appendix A.2", PCT)
    inp("DTC: sales-to-capital", 4.0, "Appendix A.2", "0.0")
    inp("DTC: investment I", OP["DTC"]["invest"], "Next-generation constellation cost (paper Sec. 5.1)")
    inp("DTC: gate year tau", OP["DTC"]["tau"], "Appendix A.2", "0")
    inp("DTC: volatility sigma", OP["DTC"]["sigma"], "Appendix A.2", "0.00")
    inp("Starship: underlying V", OP["Starship"]["V_central"], "Capitalized launch pool $14-41B/yr: CRS IF12900, NASA OIG (App. A.3)")
    inp("Starship: investment I", OP["Starship"]["invest"], "Over $15B spent plus $10-25B infrastructure (App. A.3)")
    inp("Starship: gate year tau", OP["Starship"]["tau"], "Appendix A.2", "0")
    inp("Starship: volatility sigma", OP["Starship"]["sigma"], "Appendix A.2", "0.00")
    inp("Starship: P(technical trigger)", OP["Starship"]["p_tech"], "Author judgment, varied in the inversion (paper Sec. 5.1)", PCT)

    head("Mars sovereign-program tree (paper Appendix A.2; budget analogs: ISS, Artemis, Antarctica)")
    inp("Mars: contractor margin", 0.30, "HLS-style prime-contractor margin (App. A.2)", PCT)
    inp("Mars: program start year", 10, "Appendix A.2", "0")
    inp("Mars: robotic revenue/yr", 500, "Mars budget line enacted January 2026 (App. A.3)")
    inp("Mars: P(robotic)", 0.90, "Appendix A.2", PCT)
    inp("Mars: crewed revenue/yr", 2500, "Artemis-class envelope $5-8B/yr (App. A.2)")
    inp("Mars: crewed co-investment", 5000, "HLS precedent (App. A.2)")
    inp("Mars: P(crewed given trigger)", 0.25, "Appendix A.2", PCT)
    inp("Mars: settlement revenue/yr", 10000, "Settlement-class envelope $15-30B/yr (App. A.2)")
    inp("Mars: settlement co-investment", 15000, "Appendix A.2")
    inp("Mars: P(settlement given crewed)", 0.25, "One of three crewed initiatives since 1989 survived (App. A.2)")

    head("Imported from the Python model (no formula representation)")
    inp("xAI abandonment option", round(DEC["salvage_sensitivity"]["salvage $0B"]["option_value"]),
        "Least-squares Monte Carlo (venture_option.py; paper Sec. 5.2); salvage 0")
    inp("xAI venture floor", round(DEC["venture_floor"]),
        "Venture model with optimal abandonment (paper Sec. 6 sum of parts)")
    inp("xAI winning-model value", round(DEC["xai_scenarios"]["winning model (30%)"]),
        "xAI segment at a 30% terminal margin (paper Sec. 4)")

    head("Inversion")
    inp("Target share price ($)", 135.0, "The offer price; change to read any price's implied discount rate", "0.00")


# ----------------------------------------------------------------------------------
def seg_block(ws, r0, name, rev0, tgt, m0, mT, s2c, check):
    """Classic year-by-year waterfall + step-by-step terminal value. Years in C..M."""
    cols = [col(3 + i) for i in range(T)]
    w(ws, r0, 1, name, font=H2)
    w(ws, r0, 4, "growth to 2036 target g:")
    g = f"DCF!$F${r0}"
    w(ws, r0, 6, f"=({tgt}/{rev0})^(1/{T})-1", fmt=PCT2)

    labels = ["Revenue", "(x) Operating margin", "(=) EBIT", "(-) Taxes (on positive EBIT)",
              "(-) Reinvestment (chg. revenue / sales-to-capital)", "(=) Free cash flow (FCFF)",
              "(x) Discount factor", "(=) PV of FCFF"]
    R, M_, E, X, RI, F, D, PV = (r0 + 1 + i for i in range(8))
    for i, lab in enumerate(labels):
        w(ws, r0 + 1 + i, 1, lab)
    for j, c in enumerate(cols):
        tref = f"{c}$4"
        prev = f"{cols[j - 1]}{R}" if j else f"({rev0})"
        w(ws, R, 3 + j, f"={prev}*(1+{g})", fmt=M0)
        w(ws, M_, 3 + j, f"={m0}+({mT}-{m0})*{tref}/{T}", fmt=PCT)
        w(ws, E, 3 + j, f"={c}{R}*{c}{M_}", fmt=M0)
        w(ws, X, 3 + j, f"={c}$5*MAX({c}{E},0)", fmt=M0)
        w(ws, RI, 3 + j, f"=MAX({c}{R}-{prev},0)/{s2c}", fmt=M0)
        w(ws, F, 3 + j, f"={c}{E}-{c}{X}-{c}{RI}", fmt=M0)
        w(ws, D, 3 + j, f"={c}$6", fmt="0.000")
        w(ws, PV, 3 + j, f"={c}{F}*{c}{D}", fmt=M0)

    # terminal value, step by step
    t0 = PV + 2
    steps = [
        ("Terminal value (2037 onward)", None, None),
        ("EBIT 2037 after tax = EBIT 2036 x (1+g) x (1-marginal tax)",
         f"=M{E}*(1+{IN['Terminal growth']})*(1-{IN['Tax rate, marginal']})", M0),
        ("(-) Perpetuity reinvestment: g / ROIC of the above",
         f"=B{t0 + 1}*{IN['Terminal growth']}/{IN['Terminal ROIC']}", M0),
        ("(=) FCFF 2037", f"=B{t0 + 1}-B{t0 + 2}", M0),
        ("Terminal value at 2036 = FCFF 2037 / (WACC - g)",
         f"=B{t0 + 3}/({IN['Cost of capital (WACC)']}-{IN['Terminal growth']})", M0),
        ("PV of terminal value = TV x discount factor 2036", f"=B{t0 + 4}*M{D}", M0),
    ]
    w(ws, t0, 1, steps[0][0], font=H2)
    for i, (lab, f, fmt) in enumerate(steps[1:], start=1):
        w(ws, t0 + i, 1, lab)
        w(ws, t0 + i, 2, f, fmt=fmt)
    v = t0 + 6
    w(ws, v, 1, "Segment operating value = sum of PV(FCFF) + PV(TV)", bold=True)
    w(ws, v, 2, f"=SUM(C{PV}:M{PV})+B{t0 + 5}", fmt=M0, bold=True)
    w(ws, v, 3, "Python check:", font=GREY)
    w(ws, v, 4, round(check), font=GREY, fmt=M0)

    DCFREF[name] = f"DCF!$B${v}"
    DCFREF[name + "_FCFF_row"] = F
    DCFREF[name + "_FCFF2037"] = f"DCF!$B${t0 + 3}"
    DCFREF[name + "_g"] = g
    DCFREF[name + "_margin_row"] = M_
    return v + 2


def build_dcf(ws):
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 13
    for i in range(T):
        ws.column_dimensions[col(3 + i)].width = 10
    w(ws, 1, 1, "Deterministic DCF — paper Section 4, Appendix A.1 (all values $M)", font=H1)
    w(ws, 2, 1, "The classic waterfall, year by year: revenue, margin, EBIT, taxes, reinvestment, "
                "FCFF, discounting; then the terminal value step by step. Margins and the tax rate "
                "ramp linearly; taxes apply to positive EBIT only; reinvestment is floored at zero.",
      font=GREY)
    w(ws, 4, 1, "Year index t")
    w(ws, 5, 1, "Tax rate (ramps to marginal)")
    w(ws, 6, 1, "Discount factor 1/(1+WACC)^t")
    for j in range(T):
        c = col(3 + j)
        w(ws, 3, 3 + j, str(2026 + j), font=GREY)
        w(ws, 4, 3 + j, j + 1, fmt="0")
        w(ws, 5, 3 + j, f"={IN['Tax rate, current effective']}+({IN['Tax rate, marginal']}-"
                        f"{IN['Tax rate, current effective']})*{c}$4/{T}", fmt=PCT)
        w(ws, 6, 3 + j, f"=1/(1+{IN['Cost of capital (WACC)']})^{c}$4", fmt="0.000")

    seg = DEC["segments"]
    r = 8
    r = seg_block(ws, r, "Launch", IN["Launch: revenue 2025"], IN["Launch: revenue target 2036"],
                  IN["Launch: operating margin 2025"], IN["Launch: terminal margin"],
                  IN["Launch: sales-to-capital"], seg["Launch"])
    r = seg_block(ws, r, "Starlink", IN["Starlink: revenue 2025"], IN["Starlink: revenue target 2036"],
                  IN["Starlink: operating margin 2025"], IN["Starlink: terminal margin"],
                  IN["Starlink: sales-to-capital"], seg["Starlink"])
    r = seg_block(ws, r, "xAI", IN["xAI: revenue 2025"], IN["xAI: revenue target 2036"],
                  IN["xAI: operating margin 2025"], IN["xAI: terminal margin"],
                  IN["xAI: sales-to-capital"], seg["xAI"])
    r = seg_block(ws, r, "DTC underlying", IN["DTC: 2025 revenue"], IN["DTC: revenue target 2036"],
                  IN["DTC: margin 2025"], IN["DTC: terminal margin"], IN["DTC: sales-to-capital"],
                  OP["DTC"]["E_VDTC"])

    w(ws, r, 1, "Equity bridge", font=H2)
    rows = [("Launch operating value", DCFREF["Launch"]), ("Starlink operating value", DCFREF["Starlink"]),
            ("xAI operating value", DCFREF["xAI"]), ("(+) Cash", IN["Cash"]),
            ("(+) IPO proceeds", IN["IPO proceeds"]), ("(-) Debt", IN["Debt"])]
    for i, (lab, ref) in enumerate(rows):
        w(ws, r + 1 + i, 1, lab)
        sign = "-" if "Debt" in lab else ""
        w(ws, r + 1 + i, 2, f"={sign}{ref}", font=GREEN, fmt=M0)
    w(ws, r + 7, 1, "Equity value ($M)", bold=True)
    w(ws, r + 7, 2, f"=SUM(B{r + 1}:B{r + 6})", fmt=M0, bold=True)
    w(ws, r + 7, 3, "Python check:", font=GREY)
    w(ws, r + 7, 4, round(DEC["segments"]["Launch"] + DEC["segments"]["Starlink"]
                          + DEC["segments"]["xAI"] + DEC["net_cash"]), font=GREY, fmt=M0)
    w(ws, r + 8, 1, "Per share ($)", bold=True)
    w(ws, r + 8, 2, f"=B{r + 7}/{IN['Shares outstanding (M)']}", fmt="0.00", bold=True)
    DCFREF["Equity"] = f"DCF!$B${r + 7}"


# ----------------------------------------------------------------------------------
def build_options(ws):
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 95
    w(ws, 1, 1, "Real options — paper Section 5, Appendix A.2 (values $M)", font=H1)
    w(ws, 2, 1, "Expansion claims by their standalone closed form p x [M N(d1) - I N(d2)] x (1+r)^-tau, "
                "validated against the paper's simulation in Appendix A.2. Mars as the exact "
                "probability tree. The abandonment option is imported from the Python "
                "least-squares Monte Carlo.", font=GREY)
    wacc = IN["Cost of capital (WACC)"]

    def gate(r0, title, Mref, Iref, tref, sref, pref, check_closed, check_sim):
        w(ws, r0, 1, title, font=H2)
        w(ws, r0 + 1, 1, "Underlying M")
        w(ws, r0 + 1, 2, f"={Mref}", font=GREEN, fmt=M0)
        w(ws, r0 + 2, 1, "d1 = [ln(M/I) + sigma^2/2] / sigma")
        w(ws, r0 + 2, 2, f"=(LN(B{r0 + 1}/{Iref})+0.5*{sref}^2)/{sref}", fmt="0.000")
        w(ws, r0 + 3, 1, "d2 = d1 - sigma")
        w(ws, r0 + 3, 2, f"=B{r0 + 2}-{sref}", fmt="0.000")
        w(ws, r0 + 4, 1, "Option value")
        w(ws, r0 + 4, 2, f"={pref}*(B{r0 + 1}*_xlfn.NORM.S.DIST(B{r0 + 2},TRUE)-{Iref}*"
                         f"_xlfn.NORM.S.DIST(B{r0 + 3},TRUE))/(1+{wacc})^{tref}", fmt=M0, bold=True)
        w(ws, r0 + 4, 3, round(check_closed), font=GREY, fmt=M0)
        w(ws, r0 + 4, 4, f"Python checks: closed form {check_closed:,.0f}; the paper's simulated value "
                         f"{check_sim:,.0f} (same paths as the firm model; gap = Monte-Carlo error)",
          font=GREY)
        return f"Options!$B${r0 + 4}"

    OPTREF["DTC"] = gate(4, "Starlink direct-to-cell (exercise-at-gate)",
                         DCFREF["DTC underlying"], IN["DTC: investment I"], IN["DTC: gate year tau"],
                         IN["DTC: volatility sigma"], "1",
                         DEC["options_closed_form"]["DTC"], DEC["options"]["DTC"])
    OPTREF["Starship"] = gate(10, "Starship heavy lift (gated on the technical trigger)",
                              IN["Starship: underlying V"], IN["Starship: investment I"],
                              IN["Starship: gate year tau"], IN["Starship: volatility sigma"],
                              IN["Starship: P(technical trigger)"],
                              DEC["options_closed_form"]["Starship"], DEC["options"]["Starship"])

    r = 16
    w(ws, r, 1, "Mars sovereign-program claim (exact probability tree)", font=H2)
    w(ws, r + 1, 1, "State (highest realized)")
    w(ws, r + 1, 2, "Value ($M)", bold=True)
    w(ws, r + 1, 3, "Probability", bold=True)
    tree = [
        ("Settlement", IN["Mars: settlement revenue/yr"], IN["Mars: settlement co-investment"],
         f"={IN['Starship: P(technical trigger)']}*{IN['Mars: P(crewed given trigger)']}*{IN['Mars: P(settlement given crewed)']}"),
        ("Crewed, not settlement", IN["Mars: crewed revenue/yr"], IN["Mars: crewed co-investment"],
         f"={IN['Starship: P(technical trigger)']}*{IN['Mars: P(crewed given trigger)']}*(1-{IN['Mars: P(settlement given crewed)']})"),
        ("Robotic only", IN["Mars: robotic revenue/yr"], "0",
         f"={IN['Mars: P(robotic)']}*(1-{IN['Starship: P(technical trigger)']}*{IN['Mars: P(crewed given trigger)']})"),
    ]
    for i, (lab, rev, inv, prob) in enumerate(tree):
        rr = r + 2 + i
        w(ws, rr, 1, lab)
        w(ws, rr, 2, f"=MAX({rev}*{IN['Mars: contractor margin']}*(1-{IN['Tax rate, marginal']})"
                     f"/{wacc}-{inv},0)", fmt=M0)
        w(ws, rr, 3, prob, fmt=PCT)
    w(ws, r + 5, 1, "Mars claim, discounted from the start year", bold=True)
    w(ws, r + 5, 2, f"=SUMPRODUCT(B{r + 2}:B{r + 4},C{r + 2}:C{r + 4})/(1+{wacc})^{IN['Mars: program start year']}",
      fmt=M0, bold=True)
    w(ws, r + 5, 3, round(DEC["options"]["Mars"]), font=GREY, fmt=M0)
    w(ws, r + 5, 4, "Python check (simulated tree). Value per state = max(revenue x margin x (1-tax) "
                    "/ WACC - co-investment, 0), a flat perpetuity.", font=GREY)
    OPTREF["Mars"] = f"Options!$B${r + 5}"

    r = 23
    w(ws, r, 1, "xAI abandonment option (imported)", font=H2)
    w(ws, r + 1, 1, "Value at zero salvage")
    w(ws, r + 1, 2, f"={IN['xAI abandonment option']}", font=GREEN, fmt=M0)
    w(ws, r + 1, 4, "From the Python least-squares Monte Carlo (paper Sec. 5.2); optimal stopping has "
                    "no formula representation — see venture_option.py", font=GREY)
    r += 3
    w(ws, r, 1, "Totals", font=H2)
    w(ws, r + 1, 1, "Expansion claims + Mars")
    w(ws, r + 1, 2, f"={OPTREF['DTC']}+{OPTREF['Starship']}+{OPTREF['Mars']}", fmt=M0, bold=True)
    OPTREF["TotalExpansion"] = f"Options!$B${r + 1}"
    w(ws, r + 2, 1, "All options incl. abandonment")
    w(ws, r + 2, 2, f"=B{r + 1}+{IN['xAI abandonment option']}", fmt=M0, bold=True)
    w(ws, r + 3, 1, "Equity + all options ($M)")
    w(ws, r + 3, 2, f"={DCFREF['Equity']}+B{r + 2}", fmt=M0, bold=True)
    w(ws, r + 4, 1, "Per share ($)")
    w(ws, r + 4, 2, f"=B{r + 3}/{IN['Shares outstanding (M)']}", fmt="0.00", bold=True)


# ----------------------------------------------------------------------------------
def build_montecarlo(ws):
    ws.column_dimensions["A"].width = 30
    w(ws, 1, 1, f"Monte Carlo — the distribution of equity value ({NTRIALS:,} trials; paper Sec. 4.3)",
      font=H1)
    w(ws, 2, 1, "Each row is one trial of the correlated three-segment model of Appendix A.1: a common "
                "demand factor per year (correlation rho), declining revenue volatility, the same "
                "cash-flow construction as the DCF tab. Press F9 to redraw. Excel's RAND() is not "
                "seeded, so figures move by Monte-Carlo error around the Python values; the paper's "
                "numbers use 200,000 fixed-seed paths.", font=GREY)

    # volatility schedule per segment, years in C..M (rows 5-7); drift row 8-10
    w(ws, 4, 1, "Year index t")
    for j in range(T):
        w(ws, 4, 3 + j, j + 1, fmt="0")
        w(ws, 3, 3 + j, str(2026 + j), font=GREY)
    for i, (name, *_rest) in enumerate(SEGS):
        rr = 5 + i
        w(ws, rr, 1, f"{name}: sigma_t")
        for j in range(T):
            c = col(3 + j)
            w(ws, rr, 3 + j,
              f"={IN[name + ': revenue volatility, initial']}*EXP(-{IN['Volatility decay kappa']}*{c}$4)"
              f"+{IN[name + ': revenue volatility, long run']}*(1-EXP(-{IN['Volatility decay kappa']}*{c}$4))",
              fmt=PCT)
    for i, (name, *_rest) in enumerate(SEGS):
        w(ws, 8 + i, 1, f"{name}: drift ln(1+g)")
        w(ws, 8 + i, 2, f"=LN(1+{DCFREF[name + '_g']})", font=GREEN, fmt="0.0000")

    # trial table
    hdr_row = 13
    first = hdr_row + 1
    last = hdr_row + NTRIALS
    base_f = 2                       # f_t in columns C..M (3..13) -> base index 2 means col 3+j
    base_z = {name: 13 + 11 * i for i, (name, *_r) in enumerate(SEGS)}   # cols 14.., 25.., 36..
    base_R = {name: 46 + 11 * i for i, (name, *_r) in enumerate(SEGS)}   # cols 47.., 58.., 69..
    base_P = {name: 79 + 11 * i for i, (name, *_r) in enumerate(SEGS)}   # PV FCFF cols 80.., 91.., 102..
    col_segval = {name: 113 + i for i, (name, *_r) in enumerate(SEGS)}   # 113,114,115 -> DI..DK
    col_equity = 117

    w(ws, hdr_row, 1, "Trial", bold=True)
    for j in range(T):
        w(ws, hdr_row, 3 + j, f"f {2026 + j}", font=GREY)
    for name, b in base_z.items():
        for j in range(T):
            w(ws, hdr_row, b + 1 + j, f"z {name[:2]} {26 + j}", font=GREY)
    for name, b in base_R.items():
        for j in range(T):
            w(ws, hdr_row, b + 1 + j, f"R {name[:2]} {26 + j}", font=GREY)
    for name, b in base_P.items():
        for j in range(T):
            w(ws, hdr_row, b + 1 + j, f"PV {name[:2]} {26 + j}", font=GREY)
    for name, c in col_segval.items():
        w(ws, hdr_row, c, f"{name} value", bold=True)
    w(ws, hdr_row, col_equity, "Equity ($M)", bold=True)

    rho = IN["Common-factor correlation rho"]
    for i in range(NTRIALS):
        r = first + i
        w(ws, r, 1, i + 1, fmt="0")
        for j in range(T):
            w(ws, r, 3 + j, "=_xlfn.NORM.S.INV(RAND())", fmt="0.00")
        for si, (name, rev0, tgt, m0, mT, s2c, v0, vbar) in enumerate(SEGS):
            bz, bR, bP = base_z[name], base_R[name], base_P[name]
            sig_row = 5 + si
            for j in range(T):
                fc = col(3 + j)
                zc = col(bz + 1 + j)
                w(ws, r, bz + 1 + j,
                  f"=SQRT({rho})*{fc}{r}+SQRT(1-{rho})*_xlfn.NORM.S.INV(RAND())", fmt="0.00")
            for j in range(T):
                Rc = col(bR + 1 + j)
                zc = col(bz + 1 + j)
                sc = col(3 + j)
                prev = f"{col(bR + j)}{r}" if j else f"{IN[name + ': revenue 2025']}"
                w(ws, r, bR + 1 + j,
                  f"={prev}*EXP($B${8 + si}-0.5*{sc}${sig_row}^2+{sc}${sig_row}*{zc}{r})", fmt=M0)
            mrow = DCFREF[name + "_margin_row"]
            for j in range(T):
                Rc = col(bR + 1 + j)
                dc = col(3 + j)
                prev = f"{col(bR + j)}{r}" if j else f"{IN[name + ': revenue 2025']}"
                w(ws, r, bP + 1 + j,
                  f"=({Rc}{r}*DCF!{dc}${mrow}"
                  f"-DCF!{dc}$5*MAX({Rc}{r}*DCF!{dc}${mrow},0)"
                  f"-MAX({Rc}{r}-{prev},0)/{IN[name + ': sales-to-capital']})*DCF!{dc}$6", fmt=M0)
            lastR = col(bR + T)
            w(ws, r, col_segval[name],
              f"=SUM({col(bP + 1)}{r}:{col(bP + T)}{r})"
              f"+{lastR}{r}*DCF!M${mrow}*(1+{IN['Terminal growth']})*(1-{IN['Tax rate, marginal']})"
              f"*(1-{IN['Terminal growth']}/{IN['Terminal ROIC']})"
              f"/({IN['Cost of capital (WACC)']}-{IN['Terminal growth']})*DCF!M$6", fmt=M0)
        segrefs = "+".join(f"{col(c)}{r}" for c in col_segval.values())
        w(ws, r, col_equity,
          f"={segrefs}+{IN['Cash']}+{IN['IPO proceeds']}-{IN['Debt']}", fmt=M0)

    # summary block (top right, columns O.. of rows 5-11 are free? use columns P..S rows 5..11)
    eqcol = col(col_equity)
    sm = [("Mean equity ($M)", f"=AVERAGE({eqcol}{first}:{eqcol}{last})",
           f"Python (200k paths): {CHK['mean']:,.0f}"),
          ("Std deviation", f"=_xlfn.STDEV.S({eqcol}{first}:{eqcol}{last})", ""),
          ("MC std error of mean", f"=_xlfn.STDEV.S({eqcol}{first}:{eqcol}{last})/SQRT({NTRIALS})", ""),
          ("5th percentile", f"=_xlfn.PERCENTILE.INC({eqcol}{first}:{eqcol}{last},0.05)",
           f"Python: {CHK['p5']:,.0f}"),
          ("Median", f"=_xlfn.PERCENTILE.INC({eqcol}{first}:{eqcol}{last},0.5)",
           f"Python: {CHK['med']:,.0f}"),
          ("95th percentile", f"=_xlfn.PERCENTILE.INC({eqcol}{first}:{eqcol}{last},0.95)",
           f"Python: {CHK['p95']:,.0f}"),
          ("P(equity > $1.8T)", f"=COUNTIF({eqcol}{first}:{eqcol}{last},\">1800000\")/{NTRIALS}",
           f"Python: {CHK['pgt']:.0%}")]
    w(ws, 4, 16, "Summary of the simulated distribution", font=H2)
    for i, (lab, f, chk) in enumerate(sm):
        w(ws, 5 + i, 16, lab)
        w(ws, 5 + i, 18, f, fmt=M0 if "P(" not in lab else PCT, bold=(i == 0))
        if chk:
            w(ws, 5 + i, 19, chk, font=GREY)

    # histogram: horizontal in the header zone (rows 5-6), clear of the trial rows (>= 14)
    w(ws, 4, 21, "Histogram ($T bins, count of trials)", font=H2)
    nb = 15
    for i in range(nb):
        lo = 200000 + i * 200000
        w(ws, 5, 21 + i, lo / 1e6, fmt="0.0")
        w(ws, 6, 21 + i,
          f"=COUNTIFS({eqcol}{first}:{eqcol}{last},\">={lo}\",{eqcol}{first}:{eqcol}{last},\"<{lo + 200000}\")",
          fmt="0")
    chart = BarChart()
    chart.type = "col"
    chart.title = "Simulated equity value ($T)"
    data = Reference(ws, min_col=21, max_col=21 + nb - 1, min_row=6, max_row=6)
    cats = Reference(ws, min_col=21, max_col=21 + nb - 1, min_row=5, max_row=5)
    chart.add_data(data, from_rows=True)
    chart.set_categories(cats)
    chart.legend = None
    chart.gapWidth = 10
    chart.height = 8
    chart.width = 15
    fix_axes(chart, yfmt="#,##0", xfmt="0.0")
    ws.add_chart(chart, "P15")


# ----------------------------------------------------------------------------------
def build_inversion(ws):
    ws.column_dimensions["A"].width = 11
    for cidx, width in [(2, 14), (3, 14), (4, 14), (5, 14), (6, 11), (8, 40), (9, 16)]:
        ws.column_dimensions[col(cidx)].width = width
    w(ws, 1, 1, "Inversion — the discount rate a price implies (paper Section 6)", font=H1)
    w(ws, 2, 1, "Cash flows held fixed (they do not depend on the WACC); the option layer held at its "
                "base value, as in the paper. Each row recomputes total value at that rate; the "
                "implied rate interpolates where value crosses the target. For more digits, use Goal "
                "Seek (Data > What-If) on the Inputs WACC cell.", font=GREY)

    w(ws, 4, 1, "Target price ($)")
    w(ws, 4, 2, f"={IN['Target share price ($)']}", font=GREEN, fmt="0.00")
    w(ws, 5, 1, "Target cap ($M)")
    w(ws, 5, 2, f"=B4*{IN['Shares outstanding (M)']}", fmt=M0)
    w(ws, 6, 1, "Option layer ($M)")
    w(ws, 6, 2, f"={OPTREF['TotalExpansion']}", font=GREEN, fmt=M0)
    w(ws, 6, 3, "held constant across rates (paper convention)", font=GREY)

    hdr = ["WACC", "PV of FCFF ($M)", "PV of TV ($M)", "Equity ($M)", "Total value ($M)", "$ / share"]
    for i, h in enumerate(hdr):
        w(ws, 8, 1 + i, h, bold=True)
    n_grid = 29
    frows = {n: DCFREF[n + "_FCFF_row"] for n, *_ in SEGS}
    A_terms = "+".join(DCFREF[n + "_FCFF2037"] for n, *_ in SEGS)
    for i in range(n_grid):
        r = 9 + i
        w(ws, r, 1, 0.05 + 0.0025 * i, fmt=PCT2)
        fcff_terms = "+".join(
            f"SUMPRODUCT(DCF!$C${fr}:$M${fr},1/(1+$A{r})^DCF!$C$4:$M$4)" for fr in frows.values())
        w(ws, r, 2, f"={fcff_terms}", fmt=M0)
        w(ws, r, 3, f"=({A_terms})/($A{r}-{IN['Terminal growth']})/(1+$A{r})^{T}", fmt=M0)
        w(ws, r, 4, f"=B{r}+C{r}+{IN['Cash']}+{IN['IPO proceeds']}-{IN['Debt']}", fmt=M0)
        w(ws, r, 5, f"=D{r}+$B$6", fmt=M0)
        w(ws, r, 6, f"=E{r}/{IN['Shares outstanding (M)']}", fmt="0.00")
    last = 9 + n_grid - 1
    w(ws, last + 2, 1, "Implied WACC", bold=True)
    w(ws, last + 2, 2, f"=INDEX(A9:A{last},MATCH($B$5,E9:E{last},-1))"
                       f"+(INDEX(E9:E{last},MATCH($B$5,E9:E{last},-1))-$B$5)"
                       f"/(INDEX(E9:E{last},MATCH($B$5,E9:E{last},-1))"
                       f"-INDEX(E9:E{last},MATCH($B$5,E9:E{last},-1)+1))*0.0025", fmt=PCT2, bold=True)
    w(ws, last + 2, 3, "Python check: 6.86% at $135", font=GREY)
    w(ws, last + 4, 1, "How the implied rate is computed: values fall as the rate rises, so MATCH"
                       " finds the last grid row where value is still above the target; INDEX reads"
                       " that row's rate and value; the formula then interpolates linearly to the"
                       " crossing between that row and the next.", font=GREY)
    w(ws, last + 3, 1, "Equity premium", bold=True)
    w(ws, last + 3, 2, f"=B{last + 2}-{IN['Terminal growth']}", fmt=PCT2, bold=True)
    w(ws, last + 3, 3, "over the 4.56% risk-free rate", font=GREY)

    w(ws, 8, 8, "One-lever implied values (paper Table 4; reproduce with Goal Seek on the Inputs tab)",
      bold=True)
    ref = [("Starlink terminal margin", "60% -> 137% (no feasible value)"),
           ("xAI terminal margin", "25% -> 86% (above every comparable)"),
           ("Launch 2036 revenue", "$40B -> $266B (the market is $14-41B/yr)"),
           ("xAI 2036 revenue", "$160B -> $640B (about 30x OpenAI)"),
           ("Starlink 2036 revenue", "$120B -> $358B (above the supported range)"),
           ("Discount rate", "8.25% -> 6.86% (inside estimation error)")]
    for i, (a, b) in enumerate(ref):
        w(ws, 9 + i, 8, a, font=GREY)
        w(ws, 9 + i, 9, b, font=GREY)

    chart = LineChart()
    chart.title = "Total value vs. discount rate"
    chart.y_axis.title = "Total value ($M)"
    chart.x_axis.title = "WACC"
    data = Reference(ws, min_col=5, min_row=8, max_row=last)
    cats = Reference(ws, min_col=1, min_row=9, max_row=last)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 16
    fix_axes(chart, yfmt="#,##0", xfmt="0.0%")
    ws.add_chart(chart, f"H{12 + len(ref)}")


# ----------------------------------------------------------------------------------
def build_sop(ws):
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 85
    w(ws, 1, 1, "Sum of parts against the market — paper Section 6, Table 3 ($M)", font=H1)
    rows = [
        ("Launch + Starlink (core)", f"={DCFREF['Launch']}+{DCFREF['Starlink']}", GREEN, "DCF tab"),
        ("xAI: venture floor", f"={IN['xAI venture floor']}", GREEN,
         "Low end: the venture model with optimal abandonment (imported; abandonment embedded)"),
        ("xAI: winning model", f"={IN['xAI winning-model value']}", GREEN,
         "High end: the xAI segment at a 30% terminal margin (imported)"),
        ("Net cash incl. proceeds", f"={IN['Cash']}+{IN['IPO proceeds']}-{IN['Debt']}", BLACK, ""),
        ("Expansion claims + Mars", f"={OPTREF['TotalExpansion']}", GREEN, "Options tab"),
        ("xAI abandonment option", f"={IN['xAI abandonment option']}", GREEN,
         "Added at the high end only; embedded in the venture floor at the low end"),
    ]
    for i, (lab, f, font, note) in enumerate(rows):
        w(ws, 3 + i, 1, lab)
        w(ws, 3 + i, 2, f, font=font, fmt=M0)
        w(ws, 3 + i, 3, note, font=GREY)
    w(ws, 10, 1, "Supported range: low", bold=True)
    w(ws, 10, 2, "=B3+B4+B6+B7", fmt=M0, bold=True)
    w(ws, 10, 3, f"Python check: {DEC['grounded_total_range'][0]:,.0f}", font=GREY)
    w(ws, 11, 1, "Supported range: high", bold=True)
    w(ws, 11, 2, "=B3+B5+B6+B7+B8", fmt=M0, bold=True)
    w(ws, 11, 3, f"Python check: {DEC['grounded_total_range'][1]:,.0f}", font=GREY)
    w(ws, 12, 1, "IPO capitalization", bold=True)
    w(ws, 12, 2, f"={IN['Target share price ($)']}*{IN['Shares outstanding (M)']}", fmt=M0, bold=True)
    w(ws, 13, 1, "Unexplained premium (vs. high)", bold=True)
    w(ws, 13, 2, "=B12-B11", fmt=M0, bold=True)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Components vs. the offer ($M)"
    data = Reference(ws, min_col=2, min_row=3, max_row=12)
    cats = Reference(ws, min_col=1, min_row=3, max_row=12)
    chart.add_data(data)
    chart.set_categories(cats)
    chart.legend = None
    chart.height = 9
    chart.width = 17
    fix_axes(chart, yfmt="#,##0")
    ws.add_chart(chart, "E3")


# ----------------------------------------------------------------------------------
def build_readme(ws):
    ws.column_dimensions["A"].width = 110
    lines = [
        ("SpaceX valuation — Excel companion to 'Valuing SpaceX: Cash Flows versus the Cost of Capital'", H1),
        ("Alexander F. Wagner, June 2026. Paper: ssrn.com/abstract=6918120. Code: github.com/alex-wagner-research/spacex-valuation", BLACK),
        ("", BLACK),
        ("Comments and corrections are very welcome (alexander.wagner@df.uzh.ch). This workbook was put", BLACK),
        ("together quickly around a live IPO; if something looks off, it may well be.", BLACK),
        ("", BLACK),
        ("Tabs:", H2),
        ("  Inputs      — every assumption, blue on yellow, with its source. Change these.", BLACK),
        ("  DCF         — the deterministic three-segment valuation as a classic year-by-year", BLACK),
        ("                waterfall with a step-by-step terminal value (paper Section 4), plus the", BLACK),
        ("                direct-to-cell underlying business.", BLACK),
        ("  Options     — the expansion claims by their closed forms, the Mars probability tree,", BLACK),
        ("                and the imported abandonment option (paper Section 5).", BLACK),
        ("  MonteCarlo  — 1,000 live trials of the correlated segment model (paper Section 4.3):", BLACK),
        ("                distribution summary and histogram. Press F9 to redraw.", BLACK),
        ("  Inversion   — total value across discount rates and the implied rate at the target", BLACK),
        ("                price (paper Section 6). At $135 this reads about 6.86%.", BLACK),
        ("  SumOfParts  — the supported range against the offer (paper Section 6).", BLACK),
        ("", BLACK),
        ("Excel is all you need: change the blue inputs and every tab recalculates live. The only", BLACK),
        ("cells that do NOT recalculate are the three imported xAI values on the Inputs tab", BLACK),
        ("(abandonment option, venture floor, winning-model value): they come from the Python model,", BLACK),
        ("so after large changes to the xAI assumptions, treat them as stale.", BLACK),
        ("", BLACK),
        ("Colors: blue = inputs; black = formulas; green = links across tabs; grey italics = the", BLACK),
        ("Python pipeline's values, quoted for checking. The Monte Carlo uses Excel's unseeded RAND(),", BLACK),
        ("so its figures move by sampling error around the Python values (which use 200,000 fixed-seed", BLACK),
        ("paths).", BLACK),
        ("", BLACK),
        ("Not in this workbook (Python only): the simulated option values on correlated paths (the", BLACK),
        ("closed forms here are validated against them in Appendix A.2), the xAI abandonment option's", BLACK),
        ("least-squares Monte Carlo (imported on the Inputs tab), and the joint acceptance sampling of", BLACK),
        ("Appendix B.", BLACK),
    ]
    for i, (txt, font) in enumerate(lines):
        w(ws, 1 + i, 1, txt, font=font)


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "README"
    build_readme(ws)
    build_inputs(wb.create_sheet("Inputs"))
    build_dcf(wb.create_sheet("DCF"))
    build_options(wb.create_sheet("Options"))
    build_montecarlo(wb.create_sheet("MonteCarlo"))
    build_inversion(wb.create_sheet("Inversion"))
    build_sop(wb.create_sheet("SumOfParts"))
    wb.save(OUT)
    print("Workbook written:", OUT)


if __name__ == "__main__":
    main()
